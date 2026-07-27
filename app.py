import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import re
import zipfile
import tempfile
import json
import time
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from github import Github

# ==========================================
# CONFIG & UTILS GLOBAL
# ==========================================
st.set_page_config(page_title="CCF Report Generator", page_icon="📊", layout="wide")

MONTH_ID = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

# Mapping Default (Bawaan)
DEFAULT_MAPPING = {
    'DED-ADR-ADR-SKB-CCF': 'Adrian Hidayat - Cahya Sukabumi',
    'DED-ADR-ADR-JKT-CCF': 'Adrian Hidayat - Deni Jakarta',
    'DED-ADR-IML-JKT-CCF': 'Adrian Hidayat - Imelda Jakarta',
    'DED-AGS-KDN-JKT-CCF': 'Agung Santika - Jakarta',
    'DED-DED-RUT-IDP-CCF': 'Independent - IB Semarang 2',
    'DED-STY-AYN-BJM-CCF': 'Setyo - Banjarmasin',
    'DED-SKT-TIR-BDG-CCF': 'Skot - Bandung Tera',
    'BUD-BUD-JKT-JKT-CCF': 'Budi Utama - Jakarta',
    'HAN-HAN-JKT-JKT-CCF': 'Handy - Jakarta',
    'LOY-LOY-BDG-BDG-CCF': 'Loy - Bandung',
    'OLL-OLL-SBY-SBY-CCF': 'Ollie - Surabaya',
    'REN-REN-SBY-SBY-CCF': 'Rendy Tito - Surabaya',
    'SAI-SAI-MAK-MAK-CCF': 'Saiful - Makassar',
    'SAI-LIL-RYL-MDN-CCF': 'Saiful William - Medan',
    'DED-DED-HTD-IDP-CCF': 'Hartadi - ',
    'DED-RUT-IMP-JKT-CCF': 'Independent - ',
    'SAI-DON-MAK-MAK-CCF': 'Saiful - Makassar',
    'OLL-OLL-PKB-PKB-CCF': 'Ollie - Pekanbaru',
    'OLL-OLL-SMG-SMG-CCF': 'Ollie - Semarang',
    'ALI-ALI-BDG-BDG-CCF': 'Ali Imron - Bandung',
    'DED-ALI-BDG-BDG-CCF': 'Ali Imron - Bandung',
    'DED-ASP-IDP-IDP-CCF': 'Asep - IB ',
    'DED-KWO-KWO-SBY-CCF': 'Kwok Yong - Surabaya',
    'DED-RIC-SMG-SMG-CCF': 'Richo - Semarang',
    'DED-DED-YUD-IDP-CCF': 'Yudi Januar - Yazid Januar',
    'DED-HTD-REZ-JKT-CCF': 'Reza - Jakarta',
    'DED-SKT-BAL-BAL-CCF': 'Skot - Bali',
    'DED-HTD-MUL-YOG-CCF': 'Hartadi - Jogja',
    'IDP-IDP-JKT-JKT-CCF': ' - Jakarta',
    'IDP-BOK-JKT-JKT-CCF': 'Independent Kuswan Bok - Jakarta',
    'DED-LIN-BDG-BDG-CCF': 'Lina Marlina - Bandung'
}

MAPPING_FILE = "leader_mapping.json"

def load_mapping():
    """Membaca mapping leader dari file JSON lokal, jika tidak gunakan default."""
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, "r") as f:
                return json.load(f)
        except:
            pass
    return DEFAULT_MAPPING.copy()

def save_mapping(mapping_dict):
    """Menyimpan data mapping ke file lokal, lalu Push permanen ke GitHub."""
    # 1. Simpan ke sistem lokal
    with open(MAPPING_FILE, "w") as f:
        json.dump(mapping_dict, f, indent=4)
        
    # 2. Push ke GitHub (Permanen)
    gh_token = st.secrets.get("GITHUB_TOKEN")
    gh_repo_name = st.secrets.get("GITHUB_REPO")
    
    if gh_token and gh_repo_name:
        try:
            g = Github(gh_token)
            repo = g.get_repo(gh_repo_name)
            content_str = json.dumps(mapping_dict, indent=4)
            commit_message = "Update leader mapping via Streamlit UI"
            
            try:
                # Jika file sudah ada di GitHub, perbarui (Update)
                contents = repo.get_contents(MAPPING_FILE)
                repo.update_file(contents.path, commit_message, content_str, contents.sha)
                st.success("☁️ Berhasil Disimpan Permanen ke GitHub (File Diperbarui)!")
            except:
                # Jika file belum ada di GitHub, buat baru (Create)
                repo.create_file(MAPPING_FILE, commit_message, content_str)
                st.success("☁️ Berhasil Disimpan Permanen ke GitHub (File Dibuat)!")
                
        except Exception as e:
            st.warning(f"⚠️ Tersimpan di lokal, tapi gagal push ke GitHub: {e}")
    else:
        st.info("ℹ️ Tersimpan di memori lokal (GITHUB_TOKEN atau GITHUB_REPO belum diatur di Secrets).")

def clean_text(value):
    if pd.isna(value): return ""
    text = str(value).replace("\xa0", " ").replace("–", "-")
    return re.sub(r"\s+", " ", text).strip()

def to_number(value):
    if pd.isna(value): return 0.0
    if isinstance(value, (int, float, np.integer, np.floating)): return float(value)
    text = clean_text(value).replace(" ", "").replace(",", "")
    if text == "" or text.lower() in ["nan", "none"]: return 0.0
    try: return float(text)
    except ValueError: return 0.0

def parse_period_label_from_summary(uploaded_file):
    uploaded_file.seek(0)
    text = uploaded_file.read().decode("utf-8", errors="ignore")
    match = re.search(r"from\s+(\d{4})\.(\d{2})\.(\d{2})\s+to\s+(\d{4})\.(\d{2})\.(\d{2})", text, flags=re.I)
    if not match: return "Periode"
    y1, m1, d1, y2, m2, d2 = map(int, match.groups())
    bulan1, bulan2 = MONTH_ID.get(m1, f"Bulan{m1:02d}"), MONTH_ID.get(m2, f"Bulan{m2:02d}")
    if y1 == y2 and m1 == m2: return f"{d1:02d} - {d2:02d} {bulan1} {y1}"
    if y1 == y2: return f"{d1:02d} {bulan1} - {d2:02d} {bulan2} {y1}"
    return f"{d1:02d} {bulan1} {y1} - {d2:02d} {bulan2} {y2}"

def read_html_report(uploaded_file):
    uploaded_file.seek(0)
    raw = pd.read_html(uploaded_file)[0]
    header_idx = None
    for i in range(len(raw)):
        row_values = [clean_text(x).lower() for x in raw.iloc[i].tolist()]
        if "login" in row_values:
            header_idx = i; break
    if header_idx is None: raise ValueError("Header 'Login' tidak ditemukan di file HTML.")
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = [clean_text(x) for x in raw.iloc[header_idx].tolist()]
    df = df.loc[:, [c for c in df.columns if c != ""]]
    df = df[df["Login"].notna()].copy()
    df["Login"] = pd.to_numeric(df["Login"], errors="coerce")
    df = df[df["Login"].notna()].copy()
    df["Login"] = df["Login"].astype(int)
    return df

def parse_charge_from_group(group_value):
    match = re.search(r"-(\d+)\s*$", clean_text(group_value))
    return int(match.group(1)) * 10 if match else 0

def map_leader_code(row):
    mapping = load_mapping()
    return mapping.get(clean_text(row.get('Leader Code', '')), clean_text(row.get('Leader', '')))

def safe_filename(value):
    text = clean_text(value) or "Tanpa Leader"
    return re.sub(r'[\\/*?:"<>|]', "_", text).strip()[:120] or "Tanpa Leader"

# ==========================================
# LOGIKA PENGOLAHAN EXCEL & REPORT
# ==========================================
def buat_excel_format_list_acc(jumlah_baris_kosong=1000):
    headers = ["Login", "Nama", "Group", "contry", "city", "addres", "id", "date", "levelrage", "Balance"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_fill = PatternFill(fill_type="solid", fgColor="00FF00")
    header_font = Font(name="Times New Roman", size=12, bold=True, color="000000")
    body_font = Font(name="Times New Roman", size=12, color="000000")
    med_side, thin_side = Side(style="medium", color="CCCCCC"), Side(style="thin", color="CCCCCC")
    header_border = Border(left=med_side, right=med_side, top=med_side, bottom=med_side)
    body_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, header_border, center_alignment

    for row in range(2, jumlah_baris_kosong + 2):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font, cell.border, cell.alignment = body_font, body_border, center_alignment

    column_widths = {"A": 12, "B": 28, "C": 24, "D": 14, "E": 12, "F": 28, "G": 10, "H": 14, "I": 14, "J": 14}
    for col_letter, width in column_widths.items(): ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for row in range(2, jumlah_baris_kosong + 2):
        ws.cell(row=row, column=1).number_format = "0"
        ws.cell(row=row, column=7).number_format = "0"
        ws.cell(row=row, column=10).number_format = "#,##0.00"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def build_sheet_df(base_df, include_total=True):
    result = pd.DataFrame({
        "Login": base_df["Login"].astype(int),
        "Nama": base_df["Nama"].map(clean_text),
        "Deposit": base_df["Deposit_Final"].round(2),
        "Withdrawl": base_df["Withdraw_Final"].round(2),
        "In/Out": (base_df["Deposit_Final"] + base_df["Withdraw_Final"]).round(2),
        "Total Lot": base_df["Volume"].round(2),
        "Total Komisi ": base_df["Commission"].round(2),
        "Charge Komisi": base_df["Charge Komisi"].astype(int),
        "Rate": base_df["Rate"],
        "Leader": base_df["Leader"].map(clean_text),
        "Leader Code": base_df["Leader Code"].map(clean_text),
        "Grup Meta Manager": base_df["Grup Meta Manager"].map(clean_text),
    })

    if not include_total: return result.reset_index(drop=True)

    total = {
        "Login": len(result), "Nama": "TOTAL",
        "Deposit": result["Deposit"].sum().round(2),
        "Withdrawl": result["Withdrawl"].sum().round(2),
        "In/Out": result["In/Out"].sum().round(2),
        "Total Lot": result["Total Lot"].sum().round(2),
        "Total Komisi ": result["Total Komisi "].sum().round(2),
        "Charge Komisi": "", "Rate": "", "Leader": "", "Leader Code": "", "Grup Meta Manager": "",
    }
    return pd.concat([result, pd.DataFrame([total])], ignore_index=True)

def format_workbook(output_path):
    wb = load_workbook(output_path)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    currency_fmt = '$#,##0.00;-$#,##0.00;$0.00'

    for ws in wb.worksheets:
        max_row, max_col = ws.max_row, ws.max_column
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row == 1: cell.font = Font(bold=True)

        for col in [3, 4, 5, 7]:
            for r in range(2, max_row + 1): ws.cell(r, col).number_format = currency_fmt
        for r in range(2, max_row + 1):
            ws.cell(r, 1).number_format = '0'
            ws.cell(r, 6).number_format = '#,##0.00'
            ws.cell(r, 8).number_format = '0'
            ws.cell(r, 9).number_format = '0'

        widths = {"A":11, "B":28, "C":14, "D":14, "E":14, "F":12, "G":14, "H":14, "I":12, "J":34, "K":24, "L":22}
        for col_letter, width in widths.items(): ws.column_dimensions[col_letter].width = width
    wb.save(output_path)

def create_per_leader_files(combined_df, period_label, out_dir):
    df_lite = combined_df.copy()
    df_lite["Leader"] = df_lite["Leader"].map(clean_text)
    df_lite = df_lite[df_lite["Leader"] != ""].copy()
    df_lite["__KategoriFile__"] = np.where(df_lite["Leader Code"].map(clean_text).str.upper().str.startswith("DED"), "DED", "Umum")

    def excel_round(value, digits=0):
        val = Decimal(str(to_number(value)))
        if digits >= 0: return float(val.quantize(Decimal("1").scaleb(-digits), rounding=ROUND_HALF_UP))
        factor = Decimal("1").scaleb(-digits)
        return float((val / factor).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * factor)

    def rupiah_round(value, rate): return int(excel_round(to_number(value) * to_number(rate), -3))

    created = []
    for (kategori, leader), grp in df_lite.groupby(["__KategoriFile__", "Leader"], dropna=True):
        safe_leader = safe_filename(leader)
        file_base = f"Lot-Inout {period_label}" if kategori == "DED" else f"Lot-Inout Pak Andhy {period_label}"
        out_file = Path(out_dir) / f"{file_base} ({safe_leader}).xlsx"
        sheet_name = safe_leader[:31] or "Sheet1"

        to_write = pd.DataFrame({
            "Login": grp["Login"].astype(int),
            "Nama": grp["Nama"].map(clean_text),
            "Deposit": [rupiah_round(v, r) for v, r in zip(grp["Deposit"], grp["Rate"])],
            "Withdrawl": [rupiah_round(v, r) for v, r in zip(grp["Withdrawl"], grp["Rate"])],
            "In/Out": [rupiah_round(v, r) for v, r in zip(grp["In/Out"], grp["Rate"])],
            "Total Lot": grp["Total Lot"].round(2),
            "Rate": grp["Rate"].map(to_number).astype(int),
            "Leader": grp["Leader"].map(clean_text),
        }).sort_values(["Rate", "Login"], kind="stable").reset_index(drop=True)

        total_row = {
            "Login": "TOTAL", "Nama": "",
            "Deposit": int(to_write["Deposit"].sum()), "Withdrawl": int(to_write["Withdrawl"].sum()),
            "In/Out": int(to_write["In/Out"].sum()), "Total Lot": round(float(to_write["Total Lot"].sum()), 2),
            "Rate": "", "Leader": "",
        }
        to_write = pd.concat([to_write, pd.DataFrame([total_row])], ignore_index=True)

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            to_write.to_excel(writer, index=False, sheet_name=sheet_name)

        wb = load_workbook(out_file)
        ws = wb[sheet_name]
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        rupiah_fmt = '_-"Rp"* #,##0.00_-;\\-"Rp"* #,##0.00_-;_-"Rp"* "-"??_-;_-@_-'
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row == 1 or cell.row == ws.max_row: cell.font = Font(bold=True)

        for r in range(2, ws.max_row + 1):
            ws.cell(r, 1).number_format = '0'
            for col in [3, 4, 5]: ws.cell(r, col).number_format = rupiah_fmt
            ws.cell(r, 6).number_format = '#,##0.00'
            ws.cell(r, 7).number_format = '0'

        widths = {"A":7, "B":31, "C":19, "D":18, "E":18, "F":9, "G":6, "H":19}
        for col_letter, width in widths.items(): ws.column_dimensions[col_letter].width = width

        wb.save(out_file)
        created.append(out_file)
    return created

def export_single_sheet_file(df_out, output_path, sheet_name):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(output_path)


# ==========================================
# ANTARMUKA STREAMLIT
# ==========================================
st.title("📑 Generator Laporan Komisi & ACC")

tab1, tab2, tab3 = st.tabs(["📝 Tugas 1: Buat Template List ACC", "📁 Tugas 2: Generate Report CCF (ZIP)", "⚙️ Tugas 3: Pengaturan Data Leader"])

# ----------------- TAB 1 -----------------
with tab1:
    st.header("Buat Template Excel List ACC")
    st.markdown("Fitur ini membuat file Excel kosong dengan format standar (Header hijau, border, format angka) untuk List ACC.")
    
    jml_baris = st.number_input("Jumlah baris kosong yang disiapkan:", min_value=10, max_value=5000, value=1000)
    
    if st.button("Generate Template", type="primary"):
        with st.spinner("Membuat file..."):
            excel_buffer = buat_excel_format_list_acc(jumlah_baris_kosong=jml_baris)
            st.success("✅ File berhasil dibuat!")
            st.download_button(
                label="📥 Download List ACC.xlsx",
                data=excel_buffer,
                file_name="List ACC.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ----------------- TAB 2 -----------------
with tab2:
    st.header("Generate Laporan Lot-Inout (ZIP)")
    st.markdown("Upload **ketiga** file di bawah ini untuk menghasilkan laporan final yang dikompres menjadi `.zip` otomatis.")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        f_summary = st.file_uploader("1. Summary Report (.htm/.html)", type=['htm', 'html'])
    with col2:
        f_dw = st.file_uploader("2. Deposit/Withdrawal (.htm/.html)", type=['htm', 'html'])
    with col3:
        f_acc = st.file_uploader("3. List ACC (.xlsx)", type=['xlsx', 'xls'])

    if st.button("🚀 Proses & Generate ZIP", type="primary"):
        if not (f_summary and f_dw and f_acc):
            st.error("⚠️ Silakan upload ketiga file tersebut terlebih dahulu.")
        else:
            with st.spinner("Sedang memproses laporan. Mohon tunggu..."):
                try:
                    # Buat Temp Directory
                    with tempfile.TemporaryDirectory() as tmpdir:
                        # 1. Parse File Input
                        period_label = parse_period_label_from_summary(f_summary)
                        summary = read_html_report(f_summary)
                        dw = read_html_report(f_dw)
                        
                        f_acc.seek(0)
                        accounts = pd.read_excel(f_acc)

                        # 2. Proses Summary
                        summary = summary.rename(columns={"Name": "Nama"})
                        for col in ["Deposit", "Withdraw", "In/Out", "Volume", "Commission"]:
                            if col not in summary.columns: summary[col] = 0
                            summary[col] = summary[col].map(to_number)
                        summary_keep = summary[["Login", "Volume", "Commission"]].copy()

                        # 3. Proses DW
                        dw["Comment"] = dw["Comment"].map(clean_text)
                        dw["Amount"] = dw["Amount"].map(to_number)
                        is_dep = dw["Comment"].str.lower().isin(["deposit", "margin in"])
                        is_wd = dw["Comment"].str.lower().eq("withdrawal")
                        dw_group = pd.DataFrame({
                            "Login": dw["Login"].astype(int),
                            "Deposit_Final": np.where(is_dep, dw["Amount"], 0.0),
                            "Withdraw_Final": np.where(is_wd, dw["Amount"], 0.0),
                        }).groupby("Login", as_index=False).sum()

                        # 4. Proses List ACC
                        accounts = accounts.rename(columns={"Nama":"Nama", "Group":"Grup Meta Manager", "city":"Kategori", "addres":"Leader Code", "id":"Rate"})
                        accounts["Login"] = pd.to_numeric(accounts["Login"], errors="coerce")
                        accounts = accounts[accounts["Login"].notna()].copy()
                        accounts["Login"] = accounts["Login"].astype(int)
                        accounts["Kategori"] = accounts["Kategori"].map(clean_text)
                        accounts["Leader Code"] = accounts["Leader Code"].map(clean_text)
                        accounts = accounts[accounts["Kategori"].str.upper().eq("CCF") & accounts["Leader Code"].str.upper().str.endswith("CCF")].copy()
                        
                        if "Leader" not in accounts.columns: accounts["Leader"] = ""
                        accounts["Leader"] = accounts.apply(map_leader_code, axis=1)
                        accounts["Charge Komisi"] = accounts["Grup Meta Manager"].map(parse_charge_from_group)

                        merged = accounts.merge(summary_keep, on="Login", how="left").merge(dw_group, on="Login", how="left")
                        for col in ["Volume", "Commission", "Deposit_Final", "Withdraw_Final"]:
                            merged[col] = merged[col].fillna(0).astype(float)
                        merged = merged[merged[["Deposit_Final", "Withdraw_Final", "Volume", "Commission"]].abs().sum(axis=1) > 0].copy()
                        merged = merged.sort_values(["Leader Code", "Login"], kind="stable")

                        # 5. Pisahkan DED & UMUM
                        ded_base = merged[merged["Leader Code"].str.upper().str.startswith("DED")].copy()
                        umum_base = merged[~merged["Leader Code"].str.upper().str.startswith("DED")].copy()
                        ded_df = build_sheet_df(ded_base, include_total=True)
                        umum_df = build_sheet_df(umum_base, include_total=True)
                        combined_no_total = build_sheet_df(pd.concat([ded_base, umum_base], ignore_index=True), include_total=False)

                        # 6. Buat File Output di Temp Dir
                        main_out = os.path.join(tmpdir, "CCF Hasil komisi report.xlsx")
                        ded_out = os.path.join(tmpdir, f"Laporan Lot-Inout {period_label}.xlsx")
                        umum_out = os.path.join(tmpdir, f"Laporan Lot-Inout Pak Andhy {period_label}.xlsx")
                        per_leader_dir = os.path.join(tmpdir, "per_leader_outputs")
                        
                        with pd.ExcelWriter(main_out, engine="openpyxl") as writer:
                            ded_df.to_excel(writer, sheet_name="DED", index=False)
                            umum_df.to_excel(writer, sheet_name="Umum", index=False)
                        format_workbook(main_out)

                        export_single_sheet_file(ded_df, ded_out, "DED")
                        export_single_sheet_file(umum_df, umum_out, "Umum")
                        os.makedirs(per_leader_dir, exist_ok=True)
                        created_files = create_per_leader_files(combined_no_total, period_label, per_leader_dir)

                        # 7. Zip Semuanya
                        zip_path = os.path.join(tmpdir, f"Lot-Inout {period_label}.zip")
                        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                            zf.write(main_out, arcname=os.path.basename(main_out))
                            zf.write(ded_out, arcname=f"ACT/{os.path.basename(ded_out)}")
                            zf.write(umum_out, arcname=f"ACT/{os.path.basename(umum_out)}")
                            for f in created_files:
                                zf.write(f, arcname=f"Per Leader/{os.path.basename(f)}")

                        # 8. Baca bytes ZIP untuk download
                        with open(zip_path, "rb") as f:
                            zip_bytes = f.read()

                    st.success(f"✅ Berhasil memproses data untuk periode: **{period_label}**!")
                    
                    st.download_button(
                        label="📥 Download ZIP Hasil Laporan",
                        data=zip_bytes,
                        file_name=f"Lot-Inout {period_label}.zip",
                        mime="application/zip",
                        type="primary"
                    )

                except Exception as e:
                    st.error(f"❌ Terjadi kesalahan saat memproses data: {e}")

# ----------------- TAB 3 -----------------
with tab3:
    st.header("Pengaturan Mapping Leader")
    st.markdown("Anda bisa menambah, mengedit, atau menghapus data *Leader Code* beserta Namanya langsung dari tabel di bawah ini. Pastikan untuk menekan tombol **Simpan Perubahan** setelah selesai.")

    current_mapping = load_mapping()
    
    # Ubah kamus (dictionary) menjadi DataFrame agar bisa diedit di Streamlit
    df_mapping = pd.DataFrame(list(current_mapping.items()), columns=["Leader Code", "Nama Leader"])
    
    # Render tabel yang bisa diedit (bisa tambah baris baru di bagian bawah)
    edited_df = st.data_editor(df_mapping, num_rows="dynamic", use_container_width=True, height=400)
    
    if st.button("💾 Simpan Perubahan Mapping (Permanen)", type="primary"):
        # Ubah kembali DataFrame yang sudah diedit menjadi kamus (dictionary)
        new_mapping = dict(zip(edited_df["Leader Code"], edited_df["Nama Leader"]))
        
        # Bersihkan data (Hapus baris yang kosong)
        new_mapping = {str(k).strip(): str(v).strip() for k, v in new_mapping.items() if str(k).strip() != "" and str(k).strip() != "None"}
        
        # Eksekusi simpan & push ke Github
        with st.spinner("Menyimpan ke GitHub..."):
            save_mapping(new_mapping)
    
    st.divider()
    st.subheader("Sistem Keamanan & Backup")
    st.info("Sebagai cadangan, Anda bisa mendownload file mapping saat ini, atau memulihkan data lama jika diperlukan.")
    
    colA, colB = st.columns(2)
    with colA:
        json_data = json.dumps(current_mapping, indent=4)
        st.download_button(
            label="📥 Download Backup Mapping (.json)",
            data=json_data,
            file_name="leader_mapping.json",
            mime="application/json"
        )
    
    with colB:
        uploaded_backup = st.file_uploader("📂 Restore dari File Backup (.json)", type=["json"])
        if uploaded_backup is not None:
            if st.button("🔄 Pulihkan & Push dari File"):
                try:
                    restored_mapping = json.load(uploaded_backup)
                    with st.spinner("Memulihkan ke GitHub..."):
                        save_mapping(restored_mapping)
                    st.success("✅ Data berhasil dipulihkan! Halaman akan dimuat ulang...")
                    time.sleep(2)
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal memulihkan file: {e}")